"""
Write the member reference spreadsheet straight from Memberstack.

The spreadsheet Mike shares with the membership secretary has to match what
the website actually enforces, so this reads the live member list rather than
the import CSV. Read-only against Memberstack: it never creates, updates or
removes anything.

Three sections, in the order that answers the questions people actually ask:
  1. Board members, open ended access ("never" in accessexpiresat)
  2. Everyone else, soonest expiry first
  3. Anyone still held back, read from the import CSV, because by definition
     they have no Memberstack record to read

Memberstack stores accessexpiresat as DD/MM/YYYY text, so its own dashboard
sorts that column by day of month. Section 2 is the fix for that.

The output filename is deliberately fixed and carries no date: it is shared
from iCloud, and a new filename each time would break the share and leave
whoever it was shared with looking at a stale copy. The date the export ran
goes inside the sheet instead.

Overwrites the output file every run. Nothing in it is meant to be edited by
hand: fix the data in Memberstack, then re-export.

Required environment variables:
  MEMBERSTACK_SECRET_KEY   - Memberstack Admin API secret key, Live mode
  MANUAL_ACCESS_PLAN_ID    - e.g. pln_barna-member-manual-access-xx8x0ihb

Usage:
  python3 scripts/export_members_xlsx.py
  python3 scripts/export_members_xlsx.py --out _member-list/BARNA-members.xlsx
"""
import argparse
import csv
import datetime
import json
import os
import sys
import urllib.error
import urllib.request

BASE_URL = "https://admin.memberstack.com"

FIELD_FIRST = "first-name"
FIELD_LAST = "last-name"
FIELD_EXPIRY = "accessexpiresat"

# Brand colours, same as css/style.css.
NAVY = "0D2137"
TEAL = "3AABB2"
CORAL = "E8705C"
LIGHT = "D4E8EA"

DEFAULT_OUT = "_member-list/BARNA-members.xlsx"
DEFAULT_HELD = "_member-list/members-import.csv"

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit(
        "\nERROR: openpyxl is not installed, so no spreadsheet can be written.\n\n"
        "  Fix: python3 -m pip install openpyxl\n\n"
        "  Nothing was read from Memberstack.\n"
    )


def require_env(name, example=""):
    value = os.environ.get(name, "").strip()
    if not value:
        hint = f"\n  Example value: {example}" if example else ""
        sys.exit(f"\nERROR: {name} is not set, so this cannot run.{hint}\n")
    return value


API_KEY = require_env("MEMBERSTACK_SECRET_KEY")
PLAN_ID = require_env("MANUAL_ACCESS_PLAN_ID",
                      "pln_barna-member-manual-access-xx8x0ihb")

if API_KEY.startswith("sk_sb_"):
    sys.exit(
        "\nERROR: that is the TEST secret key, not the Live one.\n\n"
        "  The 'sb' after 'sk_' means sandbox, a separate account with\n"
        "  separate members. The export would come out nearly empty.\n"
    )
if not API_KEY.startswith("sk_"):
    sys.exit("\nERROR: MEMBERSTACK_SECRET_KEY should start with 'sk_'.\n")


def api(method, path):
    req = urllib.request.Request(
        f"{BASE_URL}{path}",
        headers={
            "X-API-KEY": API_KEY,
            "User-Agent": "Mozilla/5.0 (compatible; BARNA-export-script/1.0)",
            "Accept": "application/json",
        },
        method=method,
    )
    try:
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read().decode(errors="replace"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode(errors="replace").strip()
        sys.exit(f"\nERROR: {method} {path} -> HTTP {exc.code}\n\n  {detail}\n")


def fetch_all_members():
    """Every member, walking the cursor. Same loop as the other two scripts."""
    members, end_param = [], ""
    while True:
        page = api("GET", f"/members?limit=50{end_param}")
        members.extend(page["data"])
        if not page.get("hasNextPage"):
            break
        end_param = f"&after={page['endCursor']}"
    return members


def parse_uk_date(text):
    """DD/MM/YYYY only, same strictness as check_member_expiry.py."""
    try:
        day, month, year = str(text).strip().split("/")
        return datetime.date(int(year), int(month), int(day))
    except Exception:
        return None


def active_plans(member):
    return [p for p in member.get("planConnections", [])
            if p.get("status") == "ACTIVE"]


def on_manual_access(member):
    return any(p.get("planId") == PLAN_ID for p in active_plans(member))


def read_held_back(path):
    """The people with no Memberstack record, from the import CSV."""
    if not os.path.exists(path):
        return []
    with open(path, newline="", encoding="utf-8-sig") as fh:
        return [r for r in csv.DictReader(fh) if r.get("skip", "").strip()]


HEADERS = ["#", "First name", "Surname", "Email", "Access expires", "Notes"]
HEADERS_HELD = ["#", "First name", "Surname", "Email",
                "Date on the spreadsheet", "Why it needs checking"]


def build(rows_board, rows_dated, rows_held, unparsed, out_path, today):
    wb = Workbook()
    ws = wb.active
    ws.title = "Members"

    base = Font(name="Arial", size=10)
    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    state = {"row": 1}

    def cell(col, value, font=None, fill=None, border=False, centre=False):
        c = ws.cell(state["row"], col, value)
        c.font = font or base
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        if border:
            c.border = box
        if centre:
            c.alignment = Alignment(horizontal="center")
        return c

    live = len(rows_board) + len(rows_dated)
    cell(1, "BARNA membership list, as set up on the website",
         Font(name="Arial", size=14, bold=True, color=NAVY))
    state["row"] += 1
    cell(1, f"Read straight from the live member list on "
            f"{today.strftime('%d %B %Y')}. {live} members have access.",
         Font(name="Arial", size=10, italic=True, color="555555"))
    state["row"] += 2

    def section(title, subtitle, colour, headers=HEADERS):
        cell(1, title, Font(name="Arial", size=11, bold=True, color="FFFFFF"),
             fill=colour)
        for col in range(2, len(headers) + 1):
            cell(col, None, fill=colour)
        state["row"] += 1
        cell(1, subtitle, Font(name="Arial", size=9, italic=True,
                               color="555555"))
        state["row"] += 1
        for i, head in enumerate(headers, start=1):
            cell(i, head, Font(name="Arial", size=10, bold=True, color=NAVY),
                 fill=LIGHT, border=True)
        state["row"] += 1

    def write(items):
        for n, (first, last, email, date_text, note) in enumerate(items, start=1):
            for i, value in enumerate(
                    [n, first, last, email, date_text, note], start=1):
                cell(i, value, border=True, centre=i in (1, 5))
            state["row"] += 1
        state["row"] += 2

    section("1. Board members, open ended access",
            f"{len(rows_board)} people. No expiry date, they keep access "
            f"until someone removes them by hand.", NAVY)
    write(rows_board)

    section("2. Members with an expiry date, soonest first",
            f"{len(rows_dated)} people. Access ends on the date shown unless "
            f"they renew. Anyone who joined on the website renews by card "
            f"automatically and is marked as such.", TEAL)
    write(rows_dated)


    if rows_held:
        section("3. Not set up yet, dates to confirm",
                f"{len(rows_held)} people. No account and no email sent to "
                f"them. Nothing here is urgent.", CORAL, HEADERS_HELD)
        write(rows_held)

    if unparsed:
        section("4. Needs attention: date could not be read",
                f"{len(unparsed)} people. Their date is not DD/MM/YYYY, so "
                f"their access will never expire on its own. Fix in "
                f"Memberstack.", CORAL)
        write(unparsed)

    for i, width in enumerate([5, 16, 20, 34, 20, 66], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.sheet_view.showGridLines = False

    out_dir = os.path.dirname(out_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=DEFAULT_OUT,
                    help=f"where to write the spreadsheet "
                         f"(default: {DEFAULT_OUT})")
    ap.add_argument("--held", default=DEFAULT_HELD,
                    help="import CSV to read the held-back people from; "
                         "pass '' to leave that section out")
    args = ap.parse_args()

    # Everyone with any active plan, not just the legacy free one. A member who
    # joined through the website is on the paid annual plan, and an export that
    # quietly left those out would get more wrong with every signup.
    everyone = [m for m in fetch_all_members() if active_plans(m)]
    members = [m for m in everyone if on_manual_access(m)]
    paid = [m for m in everyone if not on_manual_access(m)]
    print(f"Members with access: {len(everyone)}")

    board, dated, unparsed = [], [], []
    for m in members:
        fields = m.get("customFields") or {}
        first = fields.get(FIELD_FIRST) or ""
        last = fields.get(FIELD_LAST) or ""
        email = m["auth"]["email"]
        raw = (fields.get(FIELD_EXPIRY) or "").strip()

        if raw.lower() == "never":
            board.append((first, last, email, "never", "Board, indefinite"))
        elif parse_uk_date(raw):
            # Flag the ones worth chasing now. Whoever handles renewals reads
            # this sheet, and "soonest first" alone does not say where the
            # cutoff for acting is.
            days = (parse_uk_date(raw) - datetime.date.today()).days
            if days <= 30:
                note = f"Renews in {days} days, worth chasing"
            elif days <= 60:
                note = f"Renews in {days} days"
            else:
                note = ""
            dated.append((first, last, email, raw, note))
        else:
            unparsed.append((first, last, email, raw or "(blank)",
                             "Not a DD/MM/YYYY date, so the expiry job will "
                             "never act on it"))

    board.sort(key=lambda r: (r[1].lower(), r[0].lower()))

    held = []
    for r in read_held_back(args.held) if args.held else []:
        note = r.get("review_note", "")
        if "lapsed" in note.lower():
            why = ("Looked like it had already lapsed. Not added, worth "
                   "checking whether they have renewed.")
        else:
            why = "Not added yet, the date needs confirming."
        held.append((r["first_name"], r["last_name"], r["email"],
                     r["access_expires_at"], why))
    held.sort(key=lambda r: (parse_uk_date(r[3]) or datetime.date(2100, 1, 1),
                             r[1].lower()))

    # Website members have no accessexpiresat: Stripe renews them. Show the
    # date their subscription comes up instead, a year on from joining, so they
    # sort into the same list rather than needing a section of their own. It is
    # when they will be charged, not a date access is cut off, hence the note.
    for m in paid:
        fields = m.get("customFields") or {}
        joined = datetime.date.fromisoformat((m.get("createdAt") or "")[:10])
        try:
            renews = joined.replace(year=joined.year + 1)
        except ValueError:                          # 29 February
            renews = joined.replace(year=joined.year + 1, day=28)
        dated.append((fields.get(FIELD_FIRST) or "", fields.get(FIELD_LAST) or "",
                      m["auth"]["email"], renews.strftime("%d/%m/%Y"),
                      f"Joined on the website {joined.strftime('%d/%m/%Y')}, "
                      f"renews by card automatically"))
    dated.sort(key=lambda r: (parse_uk_date(r[3]), r[1].lower()))

    today = datetime.date.today()
    build(board, dated, held, unparsed, args.out, today)

    print(f"  board, open ended: {len(board)}")
    print(f"  with an expiry date: {len(dated)} "
          f"(including {len(paid)} who joined on the website)")
    print(f"  held back, not in Memberstack: {len(held)}")
    if unparsed:
        print(f"  ⚠️  UNREADABLE DATE: {len(unparsed)} — these will never "
              f"expire on their own, see section 4")
    print()
    print(f"Wrote {args.out}")


if __name__ == "__main__":
    main()
